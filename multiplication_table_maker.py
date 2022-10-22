#!/usr/bin/env python3
# Create a program multiplicationTable.py that takes a number N from the com-
# mand line and creates an N×N multiplication table in an Excel spreadsheet.
import openpyxl
import sys
from openpyxl.styles import Font

if len(sys.argv) < 2:
    print('Usage: app.py <Table size>')
    sys.exit(1)

try:
    num = int(sys.argv[1])
except ValueError:
    print('Usage: app.py <Table size>')
    sys.exit(1)

wb = openpyxl.Workbook()
boldFont = Font(name='Calibri', size=11, bold=True)
sheet = wb.active
sheet.title = 'multiplication_table'
sheet = wb['multiplication_table']
sheet.freeze_panes = 'B2'

for row_num in range(2, num+2):
    sheet.cell(row=row_num, column=1).value = row_num - 1
    sheet.cell(row=row_num, column=1).font = boldFont
    for column_num in range(2, num+2):
        sheet.cell(row=1, column=column_num).value = column_num - 1
        sheet.cell(row=1, column=column_num).font = boldFont
        result = (sheet.cell(row=row_num, column=1).value *
                  sheet.cell(row=1, column=column_num).value)
        sheet.cell(row=row_num, column=column_num).value = result
wb.save('multiplication_table.xlsx')  # Save the workbook.
