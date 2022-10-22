#!/usr/bin/env python3
# Create a program blankRowInserter.py that takes two integers and a filename
# string as command line arguments. Let’s call the first integer N and the sec-
# ond integer M. Starting at row N, the program should insert M blank rows
# into the spreadsheet.
import openpyxl
import sys
import os


def app_usage():
    print('Usage: app.py <start_row> <rows_to_insert> <path/to/filename>')
    sys.exit(1)


if len(sys.argv) < 4:
    app_usage()

try:
    start_row = int(sys.argv[1])
except ValueError:
    app_usage()

try:
    rows_to_insert = int(sys.argv[2])
except ValueError:
    app_usage()

filename = sys.argv[3]
if not filename.endswith('xlsx'):
    app_usage()
elif not os.path.exists(filename):
    app_usage()

wb = openpyxl.load_workbook(filename)
wb_new = openpyxl.Workbook()
sheet = wb.active
sheet_new = wb_new.active
for row_num in range(1, sheet.max_row + 1):
    for column_num in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=row_num, column=column_num).value
        if row_num <= start_row:
            sheet_new.cell(row=row_num, column=column_num).value = cell_value
        else:
            sheet_new.cell(row=row_num+rows_to_insert,
                           column=column_num).value = cell_value
wb_new.save('copy_' + os.path.basename(filename))
